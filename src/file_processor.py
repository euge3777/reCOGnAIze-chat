"""
File Processing Utility for Chat Upload
Handles extraction and parsing of various file types for inclusion in chat prompts.
"""

import os
import json
import csv
import logging
from typing import Optional, Dict, List
from pathlib import Path

logger = logging.getLogger(__name__)

# Try to import optional dependencies
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class FileProcessor:
    """Utility class for processing uploaded files."""
    
    # Supported file types and their handlers
    SUPPORTED_TYPES = {
        '.txt': 'process_text_file',
        '.csv': 'process_csv_file',
        '.json': 'process_json_file',
        '.pdf': 'process_pdf_file',
        '.xlsx': 'process_excel_file',
        '.xls': 'process_excel_file',
    }
    
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB max file size
    # To avoid exceeding the model context window, we cap how much
    # extracted content from any single file is sent to the LLM.
    # ~4k characters ≈ 1000 tokens, leaving plenty of room for
    # system prompt, knowledge-base context, and chat history.
    MAX_CONTENT_CHARS = 4000
    
    @staticmethod
    def process_uploaded_file(uploaded_file) -> Optional[Dict]:
        """
        Process an uploaded Streamlit file object.
        
        Args:
            uploaded_file: Streamlit UploadedFile object
            
        Returns:
            Dictionary with file metadata and content, or None if processing fails
        """
        try:
            if uploaded_file is None:
                return None
            
            # Check file size
            file_size = len(uploaded_file.getvalue())
            if file_size > FileProcessor.MAX_FILE_SIZE:
                logger.warning(f"File {uploaded_file.name} exceeds max size ({file_size} bytes)")
                return None
            
            file_ext = Path(uploaded_file.name).suffix.lower()
            
            if file_ext not in FileProcessor.SUPPORTED_TYPES:
                logger.warning(f"Unsupported file type: {file_ext}")
                return None
            
            handler_name = FileProcessor.SUPPORTED_TYPES[file_ext]
            handler = getattr(FileProcessor, handler_name)
            
            content = handler(uploaded_file)
            
            return {
                'filename': uploaded_file.name,
                'file_type': file_ext,
                'content': content,
                'size_bytes': file_size,
            }
            
        except Exception as e:
            logger.error(f"Error processing file {uploaded_file.name}: {str(e)}")
            return None
    
    @staticmethod
    def process_text_file(uploaded_file) -> str:
        """Extract content from a text file."""
        try:
            content = uploaded_file.read().decode('utf-8')
            return content.strip()
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                uploaded_file.seek(0)
                content = uploaded_file.read().decode('latin-1')
                return content.strip()
            except Exception as e:
                logger.error(f"Failed to decode text file: {str(e)}")
                return ""
    
    @staticmethod
    def process_csv_file(uploaded_file) -> str:
        """Extract content from a CSV file and format as readable text."""
        try:
            uploaded_file.seek(0)
            text_content = uploaded_file.read().decode('utf-8')
            
            # Parse CSV
            lines = text_content.strip().split('\n')
            reader = csv.reader(lines)
            
            rows = list(reader)
            if not rows:
                return ""
            
            # Format as readable table
            formatted = "CSV Data:\n"
            formatted += "=" * 50 + "\n"
            
            # Add header
            if rows:
                header = rows[0]
                formatted += " | ".join(header) + "\n"
                formatted += "-" * 50 + "\n"
                
                # Add data rows (limit to first 100 rows)
                for row in rows[1:101]:
                    formatted += " | ".join(str(cell) for cell in row) + "\n"
                
                if len(rows) > 101:
                    formatted += f"\n... and {len(rows) - 101} more rows\n"
            
            return formatted.strip()
        except Exception as e:
            logger.error(f"Failed to process CSV file: {str(e)}")
            return ""
    
    @staticmethod
    def process_json_file(uploaded_file) -> str:
        """Extract content from a JSON file and format as readable text."""
        try:
            uploaded_file.seek(0)
            data = json.load(uploaded_file)
            
            # Pretty print with indentation
            formatted = json.dumps(data, indent=2)
            
            # Truncate if too long (keep first 5000 chars)
            if len(formatted) > 5000:
                formatted = formatted[:5000] + "\n\n[... content truncated ...]"
            
            return formatted
        except Exception as e:
            logger.error(f"Failed to process JSON file: {str(e)}")
            return ""
    
    @staticmethod
    def process_pdf_file(uploaded_file) -> str:
        """Extract text content from a PDF file."""
        if not PDF_AVAILABLE:
            return "[PDF support requires PyPDF2. Install with: pip install PyPDF2]"
        
        try:
            uploaded_file.seek(0)
            pdf_reader = PyPDF2.PdfReader(uploaded_file)
            
            text_content = ""
            # Read all pages so that downstream summarization
            # has access to the full report content.
            num_pages = len(pdf_reader.pages)
            
            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                text_content += f"\n--- Page {page_num + 1} ---\n"
                text_content += page.extract_text()

            text_content = text_content.strip()

            # More generous safeguard on length for internal
            # processing; the chat prompt will use a separate
            # summarization layer for long PDFs.
            max_pdf_chars = 20000
            if len(text_content) > max_pdf_chars:
                text_content = text_content[:max_pdf_chars] + "\n\n[... PDF content truncated for internal processing ...]"

            return text_content
        except Exception as e:
            logger.error(f"Failed to process PDF file: {str(e)}")
            return ""
    
    @staticmethod
    def process_excel_file(uploaded_file) -> str:
        """Extract content from an Excel file."""
        if not EXCEL_AVAILABLE:
            return "[Excel support requires openpyxl. Install with: pip install openpyxl]"
        
        try:
            uploaded_file.seek(0)
            workbook = openpyxl.load_workbook(uploaded_file)
            
            formatted = "Excel Data:\n"
            formatted += "=" * 50 + "\n"
            
            # Process each sheet
            for sheet_name in workbook.sheetnames[:5]:  # Limit to first 5 sheets
                sheet = workbook[sheet_name]
                formatted += f"\n--- Sheet: {sheet_name} ---\n"
                
                # Extract data (limit to first 100 rows)
                row_count = 0
                for row in sheet.iter_rows(values_only=True):
                    if row_count >= 100:
                        break
                    formatted += " | ".join(str(cell) if cell is not None else "" for cell in row) + "\n"
                    row_count += 1
                
                if sheet.max_row > 100:
                    formatted += f"... and {sheet.max_row - 100} more rows\n"
            
            if len(workbook.sheetnames) > 5:
                formatted += f"\n... and {len(workbook.sheetnames) - 5} more sheets"
            
            return formatted.strip()
        except Exception as e:
            logger.error(f"Failed to process Excel file: {str(e)}")
            return ""
    
    @staticmethod
    def format_file_content_for_prompt(file_data: Dict, include_metadata: bool = True) -> str:
        """
        Format processed file content for inclusion in LLM prompt.
        
        Args:
            file_data: Dictionary returned by process_uploaded_file()
            include_metadata: Whether to include filename and file type in output
            
        Returns:
            Formatted string for inclusion in prompt
        """
        if not file_data:
            return ""
        
        formatted = ""
        
        if include_metadata:
            formatted += f"File: {file_data['filename']} ({file_data['file_type']})\n"
            formatted += "=" * 60 + "\n"
        
        content = file_data['content'] or ""

        # Apply a global cap again at formatting time in case other
        # processors return very large strings.
        if len(content) > FileProcessor.MAX_CONTENT_CHARS:
            content = content[:FileProcessor.MAX_CONTENT_CHARS] + "\n\n[... file content truncated for length ...]"

        formatted += content
        
        if include_metadata:
            formatted += "\n" + "=" * 60 + "\n"
        
        return formatted
